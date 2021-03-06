from bs4 import BeautifulSoup
import requests
from requests.exceptions import HTTPError
import csv
import openpyxl
import numpy as np
import pandas as pd
import gensim
from gensim.models import word2vec
from nltk.corpus import stopwords
import matplotlib.pyplot as plt
import seaborn as sns





##Problem 1 & Problem 2 Combined

#Getting keywords
path = 'keywords.xlsx'
wb = openpyxl.load_workbook(path)
sheet = wb.active
keywords = []
for cell in range(2, sheet.max_column + 1):
    keywords.append(sheet.cell(1, cell).value)

#Take query and add 100 articles to csv file
def getArticles(query, writer):
    query = '+'.join(query.split(" "))
    url = 'https://www.bbc.co.uk/search?q=' + query
    page = 1
    articleLinks = [] 
    flag = False
    while len(articleLinks) != 100:
        toSearch = url + '&page=' + str(page)
        print("SEARCHING... " + toSearch)
        try:
            response = requests.get(toSearch)
            response.raise_for_status()
        except HTTPError as http_err:
            print('HTTP ERROR OCCURRED ACCESSING '+ toSearch + ': ', http_err)
            continue
        except Exception as err:
            print('OTHER ERROR OCCURRED ACCESSING '+ toSearch + ': ', err)
            continue
        else:
            text = BeautifulSoup(response.text, 'html.parser')

        for link in text.find_all('a'):
            item = link.get('href')
            if item in articleLinks:
                if item == articleLinks[0]:
                    print("BREAKING LENGTH IS ", len(articleLinks), item)
                    flag = True
                    break
                else:
                    continue
            if len(articleLinks) == 100 or page == 30:
                print("BREAKING LENGTH IS ", len(articleLinks), item)
                flag = True
                break
            elif 'bbc.co.uk/news/' in item and item[-1].isnumeric() and not('news/help-' in item):
                try:
                    linkResponse = requests.get(item)
                    linkResponse.raise_for_status()
                except HTTPError as http_err:
                    print('HTTP ERROR OCCURRED ACCESSING '+ item + ': ', http_err)
                    continue
                except Exception as err:
                    print('OTHER ERROR OCCURRED ACCESSING '+ item + ': ', err)
                    continue
                else:
                    toCheck = BeautifulSoup(linkResponse.text, 'html.parser')
                    isArticle = toCheck.find('article')
                    if isArticle:
                        articleLinks.append(item)
                        writer.writerow([query, item, isArticle.get_text(separator=' ').replace("|", " ").replace("\n", " ")])
        if flag:
            break
        page += 1
    print(articleLinks, len(articleLinks), len(set(articleLinks)))
    return articleLinks

def scrapeAll(keywords):
    fileName = 'webcontent.csv'
    with open(fileName, 'w', newline='', encoding="utf-8") as theFile:
        writer = csv.writer(theFile, delimiter='|')
        writer.writerow(["Keyword", "URL", "Content"])
        for keyword in keywords:
            getArticles(keyword, writer)
            print(keyword + " IS NOW DONE")





##Problem 3

def preprocess(df, keyword):
    text = [] 
    smallSet =  df[df['Keyword'].str.contains(keyword, case=False)]
    for article in smallSet['Content'].values.tolist():
        text.append(gensim.utils.simple_preprocess(article))
    return text

def getDistance(df, phrase1, phrase2):
    phrase1 = phrase1.lower()
    phrase2 = phrase2.lower()
    if phrase1 == phrase2:
        return 1
    #make vocab
    vocab = []
    for iWord in phrase1.split():
        vocab.extend(preprocess(df, iWord))
    for jWord in phrase2.split():
        vocab.extend(preprocess(df, jWord))
    #additional cleaning
    #remove stopwords
    filtered = []
    for word in vocab:
        if word not in stopwords.words('english'):
            filtered.append(word)
    #train model
    model = word2vec.Word2Vec(filtered, vector_size=150, window=10, min_count=2, workers=10)
    #get distance for each word 
    distance = 0
    for iWord in phrase1.split():
        value = 0
        for jWord in phrase2.split():
            value += model.wv.similarity(iWord, jWord)
        distance += value / len(phrase2.split())
    return distance / len(phrase1.split())

scrapeAll(keywords)

datafile = pd.read_csv("webcontent.csv", delimiter='|')
items = []
cols = ['Keywords']
for keyword in keywords:
    print(keyword)
    item = [keyword]
    cols.append(keyword)
    for other in keywords:
        print(other)
        distance = getDistance(datafile, keyword, other)
        item.append(distance)
    items.append(item)
distancedf = pd.DataFrame(items, columns=cols)
distancedf.to_excel("distance.xlsx", index=False)





#Problem 4

distancedf = pd.read_excel('distance.xlsx', index_col=0)

mask = np.triu(distancedf.corr())
sns.heatmap(distancedf, cmap=sns.diverging_palette(220, 20, n=200), robust=True, annot=True, annot_kws={'size':8}, cbar=True, square=True, fmt ='.3g', mask=mask)

plt.show()