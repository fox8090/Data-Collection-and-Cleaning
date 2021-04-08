from bs4 import BeautifulSoup
import requests
from requests.exceptions import HTTPError
import csv
import openpyxl
import string
import datetime
import string

import re
import multiprocessing
import pandas as pd
from collections import defaultdict
import spacy
#import logging
#logging.basicConfig(format="%(levelname)s - %(asctime)s: %(message)s", datefmt="%H:%M:%S", level=logging.INFO)
import gensim
from gensim.models import word2vec
import platform


#### Getting keywords
path = 'keywords.xlsx'
wb = openpyxl.load_workbook(path)
sheet = wb.active
keywords = []
for cell in range(2, sheet.max_column + 1):
    keywords.append(sheet.cell(1, cell).value)

#Take query and add 100 articles to csv file
def getArticles(query, writer):
    query = '+'.join(query.split(" "))
    url = 'https://www.bbc.co.uk/search?q=' + query
    page = 1
    articleLinks = [] 
    flag = False
    while len(articleLinks) != 100:
        toSearch = url + '&page=' + str(page)
        print("SEARCHING... " + toSearch)
        try:
            response = requests.get(toSearch)
            response.raise_for_status()
        except HTTPError as http_err:
            print('HTTP ERROR OCCURRED ACCESSING '+ toSearch + ': ', http_err)
            continue
        except Exception as err:
            print('OTHER ERROR OCCURRED ACCESSING '+ toSearch + ': ', err)
            continue
        else:
            text = BeautifulSoup(response.text, 'html.parser')

        for link in text.find_all('a'):
            item = link.get('href')
            if item in articleLinks:
                if item == articleLinks[0]:
                    print("BREAKING LENGTH IS ", len(articleLinks), item)
                    flag = True
                    break
                else:
                    continue
            if len(articleLinks) == 100 or page == 30:
                print("BREAKING LENGTH IS ", len(articleLinks), item)
                flag = True
                break
            elif 'bbc.co.uk/news/' in item and item[-1].isnumeric() and not('news/help-' in item):
                try:
                    linkResponse = requests.get(item)
                    linkResponse.raise_for_status()
                except HTTPError as http_err:
                    print('HTTP ERROR OCCURRED ACCESSING '+ item + ': ', http_err)
                    continue
                except Exception as err:
                    print('OTHER ERROR OCCURRED ACCESSING '+ item + ': ', err)
                    continue
                else:
                    toCheck = BeautifulSoup(linkResponse.text, 'html.parser')
                    isArticle = toCheck.find('article')
                    if isArticle:
                        articleLinks.append(item)
                        #print(isArticle.get_text(separator=' ').translate(str.maketrans('', '', string.punctuation)), item)
                        writer.writerow([query, item, isArticle.get_text(separator=' ').replace("|", " ").replace("\n", " ")])
                    #HERE MAKE GET REQUEST FOR URL. IF IT HAS AN ARTICLE TAG THEN ADD IT STRAIGHT TO CSV. I THINK
                        #write to csv            

        if flag:
            break
        page += 1
    print(articleLinks, len(articleLinks), len(set(articleLinks)))
    return articleLinks

#query = 'Advanced Persistent Threat'
#articleLinks = getArticles(query)

def scrapeAll(keywords):
    now = datetime.datetime.now()
    formatted = now.strftime("%Y-%m-%d_%H-%M-%S")
    fileName = 'keywords_'+formatted+'.csv'
    with open(fileName, 'w', newline='', encoding="utf-8") as theFile:
        writer = csv.writer(theFile, delimiter='|')
        writer.writerow(["Keyword", "URL", "Content"])
        for keyword in keywords:
            getArticles(keyword, writer)
            print(keyword + " IS NOW DONE")

#SCRAPING ALREADY DONE WITH GREAT SUCCESS BUT USE THIS TO CALL
#scrapeAll(keywords)
#exit()

datafile = pd.read_csv("keywords_2021-04-08_14-36-06.csv", delimiter='|') # need to change this
print(datafile.shape)
print(datafile.head())

def preprocess(df, keyword):
    text = [] 
    smallSet =  df[df['Keyword'].str.contains(keyword, case=False)]
    for article in smallSet['Content'].values.tolist():
        text.append(gensim.utils.simple_preprocess(article))
    return text


def getDistance(df, phrase1, phrase2):
    phrase1 = phrase1.lower()
    phrase2 = phrase2.lower()
    #make vocab
    vocab = []
    for iWord in phrase1.split():
        vocab.extend(preprocess(df, iWord))
    for jWord in phrase2.split():
        vocab.extend(preprocess(df, jWord))
    #train model
    model = word2vec.Word2Vec(vocab, vector_size=150, window=10, min_count=2, workers=10)
    #get distance for each word 
    distance = 0
    for iWord in phrase1.split():
        value = 0
        for jWord in phrase2.split():
            value += model.wv.similarity(iWord, jWord)
        distance += value / len(phrase2.split())
    print("Distance between '" + phrase1 +"' and '" + phrase2 + "' is: " + str(distance / len(phrase1.split())))
    return distance / len(phrase1.split())

getDistance(datafile, 'spy', 'attack')

'''
print(keyword)
vocab = preprocess(datafile, keyword)
model = word2vec.Word2Vec(vocab,
        vector_size=150,
        window=10,
        min_count=2,
        workers=10)

keyword = keyword.replace(' ', '_')
print(keyword)
sim = model.wv.most_similar(positive=keyword)
print("MOST LIKE ", keyword, sim)
'''        


    